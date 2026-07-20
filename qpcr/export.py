"""Excel export — export_to_excel + native-chart post-processing.

Single source of truth (extracted from the monolith). No Streamlit dependency.
"""
import io
import re
import zipfile
import logging
from typing import Dict

import numpy as np
import pandas as pd

from qpcr.analysis import AnalysisEngine


def _sanitize_sheet_name(name: str, used_names: set) -> str:
    """Sanitize Excel sheet name: remove invalid chars, truncate, deduplicate."""
    import re
    safe = re.sub(r'[\\/*\[\]:?]', '_', name)[:31]
    base = safe
    counter = 1
    while safe in used_names:
        suffix = f"_{counter}"
        safe = base[: 31 - len(suffix)] + suffix
        counter += 1
    used_names.add(safe)
    return safe


def export_to_excel(
    raw_data: pd.DataFrame,
    processed_data: Dict[str, pd.DataFrame],
    params: dict,
    mapping: dict,
    qc_stats: dict = None,
    replicate_stats: pd.DataFrame = None,
    excluded_wells=None,
    gene_display_names: dict = None,
) -> bytes:
    """Export comprehensive Excel with gene-by-gene sheets, QC report, and FC matrix.

    `gene_display_names` maps raw gene name -> user-edited display name (from Graphs
    tab). When provided, per-gene sheet names, the FC matrix index, and the chart
    Y-axis title all use the display name so Excel matches what the user sees on
    screen. Raw gene name is preserved as a `Target_Raw` column for traceability.
    """
    gene_display_names = gene_display_names or {}
    _disp = lambda g: str(gene_display_names.get(g, g))
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        # Parameters sheet
        pd.DataFrame([params]).to_excel(
            writer, sheet_name="Analysis_Parameters", index=False
        )

        # Sample mapping sheet
        pd.DataFrame([{"Original": k, **v} for k, v in mapping.items()]).to_excel(
            writer, sheet_name="Sample_Mapping", index=False
        )

        # Raw data (include mapped Condition column reflecting sample mapping)
        raw_export = raw_data.copy()
        if mapping:
            raw_export["Condition"] = raw_export["Sample"].map(
                lambda x: mapping.get(x, {}).get("condition", x)
            )
        else:
            raw_export["Condition"] = raw_export["Sample"]
        raw_export = (
            raw_export[["Well", "Sample", "Condition", "Target", "CT", "Source_File"]]
            if "Source_File" in raw_export.columns
            else raw_export
        )
        raw_export.to_excel(writer, sheet_name="Raw_Data", index=False)

        # Gene-by-gene calculations with statistical test method column
        _used_sheet_names = {"Analysis_Parameters", "Raw_Data"}
        for gene, gene_data in processed_data.items():
            display = _disp(gene)
            sheet_name = _sanitize_sheet_name(f"{display}_Analysis", _used_sheet_names)
            gene_export = gene_data.copy()
            # Preserve raw gene name so an exported sheet can be cross-referenced
            # to raw_data["Target"] even when the user has renamed the gene.
            if "Target" in gene_export.columns and display != gene:
                gene_export.insert(
                    list(gene_export.columns).index("Target") + 1,
                    "Target_Raw",
                    gene,
                )
                gene_export["Target"] = display
            if "p_value" in gene_export.columns:
                ttest_type = params.get("ttest_type", "welch")
                gene_export["Stat_Test"] = ""
                for idx, row in gene_export.iterrows():
                    n_rep = row.get("n_replicates", 0)
                    if pd.notna(row.get("p_value")) and row.get("p_value") is not np.nan:
                        if n_rep >= 2:
                            gene_export.loc[idx, "Stat_Test"] = (
                                f"{'Welch' if ttest_type == 'welch' else 'Student'} t-test (n={n_rep})"
                            )
                        elif n_rep == 1:
                            gene_export.loc[idx, "Stat_Test"] = f"One-sample t-test (n={n_rep})"
                        else:
                            gene_export.loc[idx, "Stat_Test"] = "N/A"
            gene_export.to_excel(writer, sheet_name=sheet_name, index=False)

        # Summary sheet
        if processed_data:
            non_empty = []
            for g, df in processed_data.items():
                if df.empty:
                    continue
                d = df.copy()
                if "Target" in d.columns:
                    d["Target"] = _disp(g)
                non_empty.append(d)
            if non_empty:
                all_data = pd.concat(non_empty, ignore_index=True)
                agg_dict = {"Relative_Expression": ["mean", "std", "count"]}
                if "p_value" in all_data.columns:
                    agg_dict["p_value"] = "min"
                group_cols = ["Target"]
                if "Group" in all_data.columns:
                    group_cols.append("Group")
                summary = (
                    all_data.groupby(group_cols)
                    .agg(agg_dict)
                    .round(4)
                )
                summary.to_excel(writer, sheet_name="Summary")

        # Fold Change Matrix (pivot table)
        if processed_data:
            non_empty = []
            for g, df in processed_data.items():
                if df.empty:
                    continue
                d = df.copy()
                if "Target" in d.columns:
                    d["Target"] = _disp(g)
                non_empty.append(d)
            if non_empty:
                all_data = pd.concat(non_empty, ignore_index=True)
                if "Fold_Change" in all_data.columns and "Condition" in all_data.columns:
                    fc_matrix = all_data.pivot_table(
                        values="Fold_Change",
                        index="Target",
                        columns="Condition",
                        aggfunc="first",
                    )
                    fc_matrix = fc_matrix.round(4)
                    fc_matrix.to_excel(writer, sheet_name="FC_Matrix")

        # Replicate-level fold changes
        if raw_data is not None and excluded_wells is not None:
            hk_gene = params.get("Housekeeping_Gene")
            ref_sample = params.get("Reference_Sample")
            if hk_gene and ref_sample:
                try:
                    replicate_fc = AnalysisEngine.compute_replicate_fold_changes(
                        raw_data=raw_data,
                        hk_gene=hk_gene,
                        ref_sample=ref_sample,
                        sample_mapping=mapping,
                        excluded_wells=excluded_wells,
                    )
                    if not replicate_fc.empty:
                        replicate_fc.to_excel(writer, sheet_name="Replicate_FC", index=False)
                except Exception as e:
                    import logging
                    logging.warning(f"Replicate_FC sheet skipped: {e}")

        # QC Report sheet
        _write_qc_report_sheet(writer, qc_stats, replicate_stats)

    # Post-process: add gene chart sheets with openpyxl (supports rich text axis titles)
    output = _add_gene_chart_sheets(
        output, processed_data, params, gene_display_names=gene_display_names
    )

    return output.getvalue()


def _write_qc_report_sheet(writer, qc_stats=None, replicate_stats=None):
    """Write QC Report sheet with summary stats and replicate-level CV/outlier info."""
    rows = []
    if qc_stats:
        rows.append({"Metric": "Total Wells", "Value": qc_stats.get("total_wells", "")})
        rows.append({"Metric": "Excluded Wells", "Value": qc_stats.get("excluded_wells", "")})
        rows.append({"Metric": "Active Wells", "Value": qc_stats.get("active_wells", "")})
        rows.append({"Metric": "CT Mean", "Value": qc_stats.get("ct_mean", "")})
        rows.append({"Metric": "CT SD", "Value": qc_stats.get("ct_std", "")})
        rows.append({"Metric": "CT Range", "Value": f"{qc_stats.get('ct_min', '')}-{qc_stats.get('ct_max', '')}"})
        rows.append({"Metric": "High CT Wells (>35)", "Value": qc_stats.get("high_ct_count", "")})
        rows.append({"Metric": "Low CT Wells (<10)", "Value": qc_stats.get("low_ct_count", "")})
        rows.append({"Metric": "Total Triplicates", "Value": qc_stats.get("total_triplicates", "")})
        rows.append({"Metric": "Healthy Triplicates", "Value": qc_stats.get("healthy_triplicates", "")})
        rows.append({"Metric": "Warning Triplicates", "Value": qc_stats.get("warning_triplicates", "")})
        rows.append({"Metric": "Error Triplicates", "Value": qc_stats.get("error_triplicates", "")})
        rows.append({"Metric": "Avg CV%", "Value": qc_stats.get("avg_cv_pct", "")})
        rows.append({"Metric": "Max CV%", "Value": qc_stats.get("max_cv_pct", "")})
        rows.append({"Metric": "Health Score (%)", "Value": qc_stats.get("health_score", "")})
    if rows:
        pd.DataFrame(rows).to_excel(writer, sheet_name="QC_Report", index=False, startrow=0)
    if replicate_stats is not None and not replicate_stats.empty:
        start_row = len(rows) + 3 if rows else 0
        replicate_stats.to_excel(writer, sheet_name="QC_Report", index=False, startrow=start_row)


def _add_gene_chart_sheets(output_buf, processed_data, params, gene_display_names=None):
    """Post-process Excel bytes to add per-gene chart sheets using openpyxl.

    Two-phase approach:
    Phase 1: openpyxl API for chart creation (series, errBars, axis properties)
    Phase 2: zip-level XML post-processing for elements openpyxl can't set
             (axis label fonts, rich text Y-axis title, hidden gridlines)

    `gene_display_names` maps raw gene -> display name. The sheet name and the
    Y-axis title both use the display name so they always agree.
    """
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.data_source import NumRef, NumDataSource
    from openpyxl.chart.error_bar import ErrorBars
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.utils import get_column_letter
    from lxml import etree
    import zipfile

    if not processed_data:
        return output_buf

    gene_display_names = gene_display_names or {}
    _disp = lambda g: str(gene_display_names.get(g, g))

    output_buf.seek(0)
    wb = load_workbook(output_buf)
    hk_gene = params.get("Housekeeping_Gene", params.get("reference_gene", "\u03b2-Actin"))

    _used_sheet_names = set(wb.sheetnames)
    # Track display name per chart file so phase 2 can render the right
    # Y-axis title even when sheet names were truncated/deduplicated.
    display_for_chart_order = []

    for gene, gene_data in processed_data.items():
        if gene_data is None or gene_data.empty:
            continue

        display = _disp(gene)
        display_for_chart_order.append(display)
        sheet_name = _sanitize_sheet_name(f"{display}_Chart", _used_sheet_names)
        ws = wb.create_sheet(sheet_name)

        # Build data columns
        conditions = gene_data["Condition"].tolist() if "Condition" in gene_data.columns else gene_data.get("Group", pd.Series()).tolist()
        fold_changes = gene_data.get("Fold_Change", gene_data.get("Relative_Expression", pd.Series())).tolist()
        sems = gene_data.get("SEM", pd.Series([0] * len(gene_data))).tolist()
        p_values = gene_data.get("p_value", pd.Series()).tolist()
        significances = gene_data.get("significance", pd.Series([""]*len(gene_data))).tolist()
        has_p2 = "p_value_2" in gene_data.columns
        p_values_2 = gene_data.get("p_value_2", pd.Series()).tolist() if has_p2 else []
        sig_2 = gene_data.get("significance_2", pd.Series()).tolist() if has_p2 else []
        has_p3 = "p_value_3" in gene_data.columns
        p_values_3 = gene_data.get("p_value_3", pd.Series()).tolist() if has_p3 else []
        sig_3 = gene_data.get("significance_3", pd.Series()).tolist() if has_p3 else []

        # Fold-change-domain asymmetric error bounds (Livak). Match the on-screen
        # Plotly graph instead of plotting Ct-domain SEM symmetrically on the
        # linear fold-change axis. Fall back to symmetric SEM when unavailable.
        if "FC_Error_Upper" in gene_data.columns and "FC_Error_Lower" in gene_data.columns:
            fc_err_upper = gene_data["FC_Error_Upper"].fillna(0).tolist()
            fc_err_lower = gene_data["FC_Error_Lower"].fillna(0).tolist()
        else:
            fc_err_upper = [s if pd.notna(s) else 0 for s in sems]
            fc_err_lower = list(fc_err_upper)

        n_rows = len(conditions)
        if n_rows == 0:
            continue

        hdr_row, data_start = 5, 6
        headers = ["Condition", "Fold_Change", "SEM", "p_value", "significance"]
        if has_p2:
            headers.extend(["p_value_2", "significance_2"])
        if has_p3:
            headers.extend(["p_value_3", "significance_3"])
        for ci, h in enumerate(headers):
            ws.cell(row=hdr_row, column=3 + ci, value=h)
        # Asymmetric fold-change error columns, appended after the visible data.
        err_plus_col = 3 + len(headers)
        err_minus_col = err_plus_col + 1
        ws.cell(row=hdr_row, column=err_plus_col, value="FC_Err_Upper")
        ws.cell(row=hdr_row, column=err_minus_col, value="FC_Err_Lower")
        max_cond_len = max((len(str(c)) for c in conditions), default=10)
        ws.column_dimensions["C"].width = max(18, min(max_cond_len + 2, 45))
        for col_letter in ("D", "E", "F", "G"):
            ws.column_dimensions[col_letter].width = 14

        for i in range(n_rows):
            r = data_start + i
            ws.cell(row=r, column=3, value=conditions[i])
            fc = fold_changes[i]
            ws.cell(row=r, column=4, value=fc if pd.notna(fc) else 0)
            sem = sems[i]
            ws.cell(row=r, column=5, value=sem if pd.notna(sem) else 0)
            pv = p_values[i] if i < len(p_values) else None
            ws.cell(row=r, column=6, value=pv if pd.notna(pv) else None)
            sig = significances[i] if i < len(significances) else ""
            ws.cell(row=r, column=7, value=sig if pd.notna(sig) else "")
            if has_p2:
                pv2 = p_values_2[i] if i < len(p_values_2) else None
                ws.cell(row=r, column=8, value=pv2 if pd.notna(pv2) else None)
                s2 = sig_2[i] if i < len(sig_2) else ""
                ws.cell(row=r, column=9, value=s2 if pd.notna(s2) else "")
            p3_col_offset = 10 if has_p2 else 8
            if has_p3:
                pv3 = p_values_3[i] if i < len(p_values_3) else None
                ws.cell(row=r, column=p3_col_offset, value=pv3 if pd.notna(pv3) else None)
                s3_val = sig_3[i] if i < len(sig_3) else ""
                ws.cell(row=r, column=p3_col_offset + 1, value=s3_val if pd.notna(s3_val) else "")
            eu = fc_err_upper[i] if i < len(fc_err_upper) else 0
            el = fc_err_lower[i] if i < len(fc_err_lower) else 0
            ws.cell(row=r, column=err_plus_col, value=float(eu) if pd.notna(eu) else 0)
            ws.cell(row=r, column=err_minus_col, value=float(el) if pd.notna(el) else 0)

        last_data_row = data_start + n_rows - 1

        # ---- Phase 1: Create chart via openpyxl API ----
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.varyColors = False
        chart.legend = None
        chart.title = None
        chart.width = max(15, n_rows * 1.2)
        chart.height = 7.5
        chart.gapWidth = 219
        chart.overlap = -27

        cats = Reference(ws, min_col=3, min_row=data_start, max_row=last_data_row)
        vals = Reference(ws, min_col=4, min_row=data_start, max_row=last_data_row)
        chart.add_data(vals, titles_from_data=False)
        chart.set_categories(cats)

        series = chart.series[0]

        # Bar style: white fill (bg1), black outline (tx1) 0.5pt
        bar_sp_xml = (
            '<c:spPr xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart"'
            ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
            '<a:solidFill><a:schemeClr val="bg1"/></a:solidFill>'
            '<a:ln w="6350"><a:solidFill><a:schemeClr val="tx1"/></a:solidFill>'
            '<a:prstDash val="solid"/></a:ln></c:spPr>'
        )
        series.graphicalProperties = GraphicalProperties.from_tree(etree.fromstring(bar_sp_xml))

        # Error bars: custom asymmetric fold-change bounds (upper/lower), gray line
        plus_col_letter = get_column_letter(err_plus_col)
        minus_col_letter = get_column_letter(err_minus_col)
        plus_formula = f"'{sheet_name}'!${plus_col_letter}${data_start}:${plus_col_letter}${last_data_row}"
        minus_formula = f"'{sheet_name}'!${minus_col_letter}${data_start}:${minus_col_letter}${last_data_row}"
        eb_sp_xml = (
            '<c:spPr xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart"'
            ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
            '<a:noFill/>'
            '<a:ln w="9525" cap="flat" cmpd="sng" algn="ctr">'
            '<a:solidFill><a:schemeClr val="tx1">'
            '<a:lumMod val="65000"/><a:lumOff val="35000"/>'
            '</a:schemeClr></a:solidFill>'
            '<a:prstDash val="solid"/><a:round/></a:ln></c:spPr>'
        )
        eb_sp = GraphicalProperties.from_tree(etree.fromstring(eb_sp_xml))
        series.errBars = ErrorBars(
            errValType="cust", errBarType="both", noEndCap=False,
            minus=NumDataSource(numRef=NumRef(f=minus_formula)),
            plus=NumDataSource(numRef=NumRef(f=plus_formula)),
            spPr=eb_sp,
        )

        # X-axis: black line 0.5pt, no tick marks
        chart.x_axis.delete = False
        chart.x_axis.majorTickMark = "none"
        chart.x_axis.minorTickMark = "none"
        xax_sp_xml = (
            '<c:spPr xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart"'
            ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
            '<a:noFill/>'
            '<a:ln w="6350" cap="flat" cmpd="sng" algn="ctr">'
            '<a:solidFill><a:schemeClr val="tx1"/></a:solidFill>'
            '<a:prstDash val="solid"/><a:round/></a:ln></c:spPr>'
        )
        chart.x_axis.graphicalProperties = GraphicalProperties.from_tree(etree.fromstring(xax_sp_xml))

        # Y-axis: black line 0.5pt, tick out, min=0, no gridlines
        chart.y_axis.delete = False
        chart.y_axis.majorTickMark = "out"
        chart.y_axis.minorTickMark = "none"
        chart.y_axis.scaling.min = 0
        chart.y_axis.majorGridlines = None
        yax_sp_xml = (
            '<c:spPr xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart"'
            ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
            '<a:noFill/>'
            '<a:ln w="6350"><a:solidFill><a:schemeClr val="tx1"/></a:solidFill>'
            '<a:prstDash val="solid"/></a:ln></c:spPr>'
        )
        chart.y_axis.graphicalProperties = GraphicalProperties.from_tree(etree.fromstring(yax_sp_xml))

        ws.add_chart(chart, "I3")

    # ---- Phase 1 complete: save workbook to bytes ----
    phase1_buf = io.BytesIO()
    wb.save(phase1_buf)

    # ---- Phase 2: post-process chart XMLs in the XLSX zip ----
    C = "http://schemas.openxmlformats.org/drawingml/2006/chart"
    A = "http://schemas.openxmlformats.org/drawingml/2006/main"
    nsmap = {"c": C, "a": A}

    phase1_buf.seek(0)
    zf_in = zipfile.ZipFile(phase1_buf, "r")

    # Identify which chart files are gene charts (the ones we just added).
    # Sort NUMERICALLY (chart1, chart2, ..., chart10, chart11) — a plain sorted()
    # is lexicographic (chart1, chart10, chart11, ..., chart2) which mispairs
    # genes to charts as soon as there are 10+ genes, stamping the wrong gene
    # name onto a chart's Y-axis title.
    def _chart_file_num(name):
        m = re.search(r"chart(\d+)\.xml$", name)
        return int(m.group(1)) if m else 0

    all_chart_files = sorted(
        (n for n in zf_in.namelist()
         if n.startswith("xl/charts/chart") and n.endswith(".xml")),
        key=_chart_file_num,
    )
    # Gene chart files are the last N chart files (one per gene with data), in
    # creation order. Use the display-name order captured during phase 1 so the
    # Y-axis title matches the sheet name even when gene names were renamed.
    n_gene_charts = len(display_for_chart_order)
    gene_chart_files = all_chart_files[-n_gene_charts:] if n_gene_charts > 0 else []
    gene_for_chart = dict(zip(gene_chart_files, display_for_chart_order))

    result = io.BytesIO()
    zf_out = zipfile.ZipFile(result, "w", compression=zipfile.ZIP_DEFLATED)

    for item in zf_in.infolist():
        data = zf_in.read(item.filename)

        if item.filename in gene_for_chart:
            gene = gene_for_chart[item.filename]
            root = etree.fromstring(data)

            cat_ax = root.find(".//c:catAx", nsmap)
            val_ax = root.find(".//c:valAx", nsmap)

            # X-axis: add txPr for 9pt gray Arial labels
            if cat_ax is not None:
                cat_ax.append(_build_axis_txpr(C, A))

            if val_ax is not None:
                # Hidden gridlines (noFill)
                mgrid = etree.SubElement(val_ax, f"{{{C}}}majorGridlines")
                mg_sp = etree.SubElement(mgrid, f"{{{C}}}spPr")
                mg_ln = etree.SubElement(mg_sp, f"{{{A}}}ln")
                etree.SubElement(mg_ln, f"{{{A}}}noFill")

                # Y-axis tick label font
                val_ax.append(_build_axis_txpr(C, A))

                # Rich text Y-axis title
                _build_yaxis_title(val_ax, gene, hk_gene, C, A)

            # Chart area: white fill, light gray border
            chart_space = root
            cs_sp = etree.SubElement(chart_space, f"{{{C}}}spPr")
            sf_cs = etree.SubElement(cs_sp, f"{{{A}}}solidFill")
            etree.SubElement(sf_cs, f"{{{A}}}schemeClr").set("val", "bg1")
            ln_cs = etree.SubElement(cs_sp, f"{{{A}}}ln")
            ln_cs.set("w", "9525")
            sf_ln = etree.SubElement(ln_cs, f"{{{A}}}solidFill")
            sc_ln = etree.SubElement(sf_ln, f"{{{A}}}schemeClr")
            sc_ln.set("val", "tx1")
            etree.SubElement(sc_ln, f"{{{A}}}lumMod").set("val", "15000")
            etree.SubElement(sc_ln, f"{{{A}}}lumOff").set("val", "85000")

            data = etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)

        zf_out.writestr(item, data)

    zf_in.close()
    zf_out.close()
    result.seek(0)
    return result


def _build_axis_txpr(C, A):
    """Build txPr element for axis tick labels: 9pt gray Arial."""
    from lxml import etree

    txpr = etree.Element(f"{{{C}}}txPr")
    bp = etree.SubElement(txpr, f"{{{A}}}bodyPr")
    bp.set("rot", "-60000000")
    bp.set("vert", "horz")
    bp.set("wrap", "square")
    etree.SubElement(txpr, f"{{{A}}}lstStyle")
    p = etree.SubElement(txpr, f"{{{A}}}p")
    pPr = etree.SubElement(p, f"{{{A}}}pPr")
    dRPr = etree.SubElement(pPr, f"{{{A}}}defRPr")
    dRPr.set("sz", "900")
    dRPr.set("b", "0")
    sf = etree.SubElement(dRPr, f"{{{A}}}solidFill")
    sc = etree.SubElement(sf, f"{{{A}}}schemeClr")
    sc.set("val", "tx1")
    etree.SubElement(sc, f"{{{A}}}lumMod").set("val", "65000")
    etree.SubElement(sc, f"{{{A}}}lumOff").set("val", "35000")
    lat = etree.SubElement(dRPr, f"{{{A}}}latin")
    lat.set("typeface", "Arial")
    cs = etree.SubElement(dRPr, f"{{{A}}}cs")
    cs.set("typeface", "Arial")
    return txpr


def _build_yaxis_title(val_ax, gene, hk_gene, C, A):
    """Build rich text Y-axis title: 'Relative mRNA expression level of' + gene(RED)/HK(black)."""
    from lxml import etree

    title_el = etree.SubElement(val_ax, f"{{{C}}}title")
    tx = etree.SubElement(title_el, f"{{{C}}}tx")
    rich = etree.SubElement(tx, f"{{{C}}}rich")
    body = etree.SubElement(rich, f"{{{A}}}bodyPr")
    body.set("rot", "-5400000")
    body.set("vert", "horz")
    body.set("wrap", "square")
    body.set("anchor", "ctr")
    body.set("anchorCtr", "1")
    etree.SubElement(rich, f"{{{A}}}lstStyle")

    # Paragraph 1: "Relative mRNA expression level of"
    p1 = etree.SubElement(rich, f"{{{A}}}p")
    p1Pr = etree.SubElement(p1, f"{{{A}}}pPr")
    etree.SubElement(p1Pr, f"{{{A}}}defRPr")
    r1 = etree.SubElement(p1, f"{{{A}}}r")
    rPr1 = etree.SubElement(r1, f"{{{A}}}rPr")
    rPr1.set("lang", "en-US")
    rPr1.set("sz", "800")
    rPr1.set("b", "1")
    sf1 = etree.SubElement(rPr1, f"{{{A}}}solidFill")
    sc1 = etree.SubElement(sf1, f"{{{A}}}sysClr")
    sc1.set("val", "windowText")
    sc1.set("lastClr", "000000")
    lat1 = etree.SubElement(rPr1, f"{{{A}}}latin")
    lat1.set("typeface", "Arial")
    cs1 = etree.SubElement(rPr1, f"{{{A}}}cs")
    cs1.set("typeface", "Arial")
    t1 = etree.SubElement(r1, f"{{{A}}}t")
    t1.text = "Relative mRNA expression level of"

    # Paragraph 2: gene name (RED) + /HK_gene (black)
    p2 = etree.SubElement(rich, f"{{{A}}}p")
    p2Pr = etree.SubElement(p2, f"{{{A}}}pPr")
    etree.SubElement(p2Pr, f"{{{A}}}defRPr")

    # Gene name run (RED #FF0000)
    r2 = etree.SubElement(p2, f"{{{A}}}r")
    rPr2 = etree.SubElement(r2, f"{{{A}}}rPr")
    rPr2.set("lang", "en-US")
    rPr2.set("sz", "800")
    rPr2.set("b", "1")
    sf2 = etree.SubElement(rPr2, f"{{{A}}}solidFill")
    etree.SubElement(sf2, f"{{{A}}}srgbClr").set("val", "FF0000")
    lat2 = etree.SubElement(rPr2, f"{{{A}}}latin")
    lat2.set("typeface", "Arial")
    cs2 = etree.SubElement(rPr2, f"{{{A}}}cs")
    cs2.set("typeface", "Arial")
    t2 = etree.SubElement(r2, f"{{{A}}}t")
    t2.text = gene

    # /HK_gene run (black)
    r3 = etree.SubElement(p2, f"{{{A}}}r")
    rPr3 = etree.SubElement(r3, f"{{{A}}}rPr")
    rPr3.set("lang", "en-US")
    rPr3.set("sz", "800")
    rPr3.set("b", "1")
    sf3 = etree.SubElement(rPr3, f"{{{A}}}solidFill")
    sc3 = etree.SubElement(sf3, f"{{{A}}}sysClr")
    sc3.set("val", "windowText")
    sc3.set("lastClr", "000000")
    lat3 = etree.SubElement(rPr3, f"{{{A}}}latin")
    lat3.set("typeface", "Arial")
    cs3 = etree.SubElement(rPr3, f"{{{A}}}cs")
    cs3.set("typeface", "Arial")
    t3 = etree.SubElement(r3, f"{{{A}}}t")
    t3.text = f"/{hk_gene}"

    etree.SubElement(title_el, f"{{{C}}}overlay").set("val", "0")
